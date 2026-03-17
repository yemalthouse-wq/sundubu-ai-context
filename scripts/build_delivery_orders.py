#!/usr/bin/env python3
"""
build_delivery_orders.py
Uber Eats CSVとロケットナウXLSXを読み込み、
public/delivery_orders.json に統合出力するスクリプト。
"""

import csv
import json
import glob
import os
import sys
from datetime import datetime

try:
    import openpyxl
except ImportError:
    print("openpyxl is required: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
UBER_DIR = os.path.join(BASE_DIR, "data", "uber")
ROCKET_DIR = os.path.join(BASE_DIR, "data", "rocketnow")
OUTPUT_DIR = os.path.join(BASE_DIR, "public")
OUTPUT_FILE = os.path.join(OUTPUT_DIR, "delivery_orders.json")


def parse_uber_csv(filepath):
    """Uber Eats CSVファイルを読み込んで注文リストを返す。"""
    orders = []
    with open(filepath, encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            order = {
                "source": "uber_eats",
                "file": os.path.basename(filepath),
            }
            # 全カラムをそのまま取り込む
            for key, value in row.items():
                if key is None:
                    continue
                clean_key = key.strip()
                order[clean_key] = value.strip() if value else ""
            orders.append(order)
    return orders


def parse_rocketnow_xlsx(filepath):
    """ロケットナウXLSXファイルを読み込んで注文リストを返す。"""
    orders = []
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        headers = [str(h).strip() if h else f"col_{i}" for i, h in enumerate(rows[0])]
        for row in rows[1:]:
            if all(cell is None for cell in row):
                continue
            order = {
                "source": "rocketnow",
                "file": os.path.basename(filepath),
                "sheet": sheet,
            }
            for header, cell in zip(headers, row):
                if isinstance(cell, datetime):
                    order[header] = cell.isoformat()
                else:
                    order[header] = str(cell).strip() if cell is not None else ""
            orders.append(order)
    wb.close()
    return orders


def main():
    all_orders = []

    # Uber Eats CSV
    uber_files = sorted(glob.glob(os.path.join(UBER_DIR, "*.csv")))
    print(f"Uber Eats CSV files found: {len(uber_files)}")
    for fp in uber_files:
        print(f"  Reading: {fp}")
        all_orders.extend(parse_uber_csv(fp))

    # ロケットナウ XLSX
    rocket_files = sorted(glob.glob(os.path.join(ROCKET_DIR, "*.xlsx")))
    print(f"RocketNow XLSX files found: {len(rocket_files)}")
    for fp in rocket_files:
        print(f"  Reading: {fp}")
        all_orders.extend(parse_rocketnow_xlsx(fp))

    print(f"Total orders: {len(all_orders)}")

    # 出力
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    with open(OUTPUT_FILE, "w", encoding="utf-8") as f:
        json.dump(
            {
                "generated_at": datetime.now().isoformat(),
                "total_orders": len(all_orders),
                "orders": all_orders,
            },
            f,
            ensure_ascii=False,
            indent=2,
        )
    print(f"Output: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
